#!/usr/bin/env python3
"""
Step_00_lookup_from_ein.py

1) Reads the first XML and JSON for a given EIN
2) Extracts key descriptive fields (ActivityOrMissionDesc, MissionDesc, Desc)
3) Generates prioritized keywords and matches them against the NTEE lookup CSV with scoring
4) Limits to top 10 NTEE codes
5) Writes matched NTEE codes to shared/input/filter_config.json
6) Ensures a single Excel dossier per EIN (shared/output/dossiers/EIN_<EIN>_results.xlsx) and appends a record to the "Filter_Config" sheet. If the dossier file doesn't exist, it's created with the header row, then the record is appended.

Assumes Docker volume mapping: ./shared → /data inside container.
"""
from pathlib import Path
import json
import xml.etree.ElementTree as ET
import pandas as pd
import sys
import os
import re
from datetime import datetime, timezone
from openpyxl import load_workbook, Workbook

def main():
    shared_dir = Path("/data")
    print(f"[DEBUG] shared_dir   = {shared_dir}")

    xml_dir     = shared_dir / "output" / "xml"
    json_dir    = shared_dir / "output" / "990s_propublica"
    csv_path    = shared_dir / "input"  / "lookups" / "ntee_keywords.csv"
    output_json = shared_dir / "input"  / "filter_config.json"
    dossier_dir = shared_dir / "output" / "dossiers"

    # Verify required paths
    for p in [xml_dir, json_dir, csv_path, dossier_dir]:
        if not p.exists():
            print(f"ERROR: Required path not found: {p}", file=sys.stderr)
            return

    # Pick input files
    xml_files  = sorted(xml_dir.glob("*.xml"))
    json_files = sorted(json_dir.glob("*.json"))
    if not xml_files or not json_files:
        print("ERROR: No XML or JSON inputs.", file=sys.stderr)
        return
    xml_path  = xml_files[0]
    json_path = json_files[0]

    ein = xml_path.stem.split('_')[0]
    print(f"[DEBUG] EIN         = {ein}")

    # Parse XML
    ns = {'irs': 'http://www.irs.gov/efile'}
    tags = ['ActivityOrMissionDesc', 'MissionDesc', 'Desc']
    texts = []
    xml_review = 'N'
    xml_year = ''
    try:
        root = ET.parse(xml_path).getroot()
        xml_review = 'Y'
        xml_year = datetime.fromtimestamp(xml_path.stat().st_mtime, timezone.utc).year
        for tag in tags:
            el = root.find(f".//irs:{tag}", ns)
            if el is not None and el.text:
                texts.append(el.text.strip())
    except Exception as e:
        print(f"ERROR: parsing XML: {e}", file=sys.stderr)

    # Parse JSON
    json_review = 'N'
    json_year = ''
    org_name = ''
    try:
        json_data = json.loads(json_path.read_text())
        json_review = 'Y'
        json_year = datetime.fromtimestamp(json_path.stat().st_mtime, timezone.utc).year
        org_name = json_data.get('organization_name') or \
                   json_data.get('organizationName') or \
                   json_data.get('organization', {}).get('name', '')
    except Exception as e:
        print(f"ERROR: parsing JSON: {e}", file=sys.stderr)

    # Build keywords
    raw_text = " ".join(texts).lower()
    priority = [
        "health","medical","dental","mental health","free clinic",
        "low-income","underserved","uninsured","volunteer","patient care","medicaid"
    ]
    words = re.findall(r"\b[a-zA-Z\-]{4,}\b", raw_text)
    keywords = sorted(set(priority + words))
    keyword_str = ", ".join(keywords)
    print(f"[DEBUG] keywords = {keywords}")

    # Load NTEE lookup
    try:
        df = pd.read_csv(csv_path, encoding='latin1')
        df['combined'] = (df['Description'].fillna('') + ' ' + df['Definition'].fillna('')).str.lower()
    except Exception as e:
        print(f"ERROR: loading CSV: {e}", file=sys.stderr)
        return

    # Score and select top 10
    scores = {row['NTEE Code']: sum(row['combined'].count(kw) for kw in keywords) for _, row in df.iterrows()}
    scored = [(code, cnt) for code, cnt in scores.items() if cnt > 0]
    top_codes = [code for code, _ in sorted(scored, key=lambda x: x[1], reverse=True)[:10]]
    codes_str = ", ".join(top_codes)
    print(f"[DEBUG] top_codes = {top_codes}")

    # Write filter_config.json
    os.makedirs(output_json.parent, exist_ok=True)
    try:
        output_json.write_text(json.dumps({'ntee_codes': top_codes}, indent=2))
        print(f"SUCCESS: wrote {len(top_codes)} codes to {output_json}")
    except Exception as e:
        print(f"ERROR: writing config: {e}", file=sys.stderr)

    # Update or create dossier Excel
    dossier_path = dossier_dir / f"EIN_{ein}_results.xlsx"
    header = ['Date','Time','EIN','Organization Name','NTEE Codes','JSON Review','JSON Year','XML Review','XML Year','Keywords']
    if dossier_path.exists():
        try:
            wb = load_workbook(dossier_path)
        except Exception as e:
            print(f"ERROR: loading Excel dossier: {e}", file=sys.stderr)
            return
    else:
        # Create a new workbook preserving no default sheet
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        ws_init = wb.create_sheet('Filter_Config')
        ws_init.append(header)
        wb.save(dossier_path)
        wb = load_workbook(dossier_path)
    # Ensure Filter_Config sheet
    ws = wb['Filter_Config']
    # Append new record with Date, Time, and years
    now = datetime.now(timezone.utc)
    ws.append([
        now.date().isoformat(),
        now.time().strftime('%H:%M'),
        ein,
        org_name,
        codes_str,
        json_review,
        str(json_year),
        xml_review,
        str(xml_year),
        keyword_str
    ])
    try:
        wb.save(dossier_path)
        print(f"SUCCESS: updated dossier Excel at {dossier_path}")
    except Exception as e:
        print(f"ERROR: saving dossier Excel: {e}", file=sys.stderr)

if __name__=='__main__':
    main()
